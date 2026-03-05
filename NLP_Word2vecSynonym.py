import gensim.downloader as api
from gensim.models.keyedvectors import KeyedVectors
from nltk.tokenize import word_tokenize
import openpyxl

# 加载预训练的 Word2Vec 模型
model = api.load("word2vec-google-news-300")  # 这是 Google 的预训练 Word2Vec 模型

def get_synonyms_using_word2vec(word, topn=5):
    """
    使用 Word2Vec 模型获取同义词（根据语义相似度）
    :param word: 输入词
    :param topn: 返回的最相似词的数量
    :return: 返回与输入词最相似的同义词
    """
    try:
        # 获取最相似的词汇
        similar_words = model.most_similar(word, topn=topn)
        return similar_words
    except KeyError:
        # 如果模型中没有该词，返回空列表
        print(f"Word '{word}' not found in Word2Vec model.")
        return []

# 打开 Excel 文件，读取词表
input_file = 'wordlist.xlsx'
workbook = openpyxl.load_workbook(input_file)
sheet = workbook.active

# 获取词表中的所有单词（假设单词在 "word" 列）
words = []
for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):  # 跳过标题行
    words.append(row[0])

# 创建一个新的 Excel 文件用于输出同义词
output_file = 'Word2vecSynonyms.xlsx'
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active
output_sheet.append(["Original Word", "Top 1 Synonym", "Top 2 Synonym", "Top 3 Synonym", "Top 4 Synonym", "Top 5 Synonym"])

# 对每个单词，获取 Word2Vec 模型分析出的前 5 个同义词
for word in words:
    similar_words = get_synonyms_using_word2vec(word)
    synonyms = [word] + [similar_word for similar_word, _ in similar_words]  # 将原词和5个近义词合并
    # 将结果写入新的 Excel 文件
    output_sheet.append(synonyms)

# 保存结果
output_workbook.save(output_file)
print(f"Word2vec synonyms have been saved to '{output_file}'")
